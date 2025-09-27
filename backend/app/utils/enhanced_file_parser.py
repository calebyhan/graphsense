"""
Enhanced File Parser with Streaming Support and Memory Management
"""

import asyncio
import logging
from typing import Dict, Any, List, Optional, Union
from fastapi import UploadFile
import pandas as pd
import json
from io import BytesIO
import openpyxl
from openpyxl import load_workbook

from app.utils.streaming_processor import StreamingDataProcessor
from app.utils.memory_manager import get_memory_manager, RequestPriority

logger = logging.getLogger(__name__)


class EnhancedFileParser:
    """Enhanced file parser with streaming support and memory management"""
    
    def __init__(self, 
                 chunk_size: int = 10000,
                 memory_limit_mb: int = 512,
                 sample_size: int = 5000):
        """
        Initialize enhanced file parser
        
        Args:
            chunk_size: Chunk size for streaming processing
            memory_limit_mb: Memory limit for processing
            sample_size: Maximum sample size for analysis
        """
        self.streaming_processor = StreamingDataProcessor(
            chunk_size=chunk_size,
            memory_limit_mb=memory_limit_mb,
            sample_size=sample_size
        )
        self.memory_manager = get_memory_manager()
    
    async def parse_file(self, 
                        file: UploadFile, 
                        request_id: str,
                        priority: RequestPriority = RequestPriority.NORMAL) -> Dict[str, Any]:
        """
        Parse uploaded file with streaming support and memory management
        
        Args:
            file: Uploaded file
            request_id: Unique request identifier
            priority: Processing priority
            
        Returns:
            Dictionary with parsed data and metadata
        """
        try:
            # Read file content
            file_content = await file.read()
            file_size_mb = len(file_content) / (1024 * 1024)
            
            logger.info(f"Parsing file {file.filename} ({file_size_mb:.2f} MB)")
            
            # Estimate memory requirements
            estimated_memory_mb = max(50, int(file_size_mb * 2))  # Rough estimate
            
            # Determine file type
            file_extension = self._get_file_extension(file.filename)
            
            # Create parsing callback
            async def parse_callback():
                return await self._parse_file_content(
                    file_content, 
                    file.filename or "unknown", 
                    file_extension
                )
            
            # Queue the parsing request
            success = await self.memory_manager.queue_request(
                request_id=request_id,
                callback=parse_callback,
                estimated_memory_mb=estimated_memory_mb,
                priority=priority,
                timeout_seconds=300
            )
            
            if not success:
                raise Exception("Failed to queue file parsing request - system overloaded")
            
            # The callback will be executed when memory is available
            # For now, return a placeholder - in a real implementation,
            # you'd need to implement a way to wait for the result
            return await parse_callback()
            
        except Exception as e:
            logger.error(f"File parsing failed for {file.filename}: {e}")
            raise
    
    async def _parse_file_content(self, 
                                 file_content: bytes, 
                                 filename: str, 
                                 file_extension: str) -> Dict[str, Any]:
        """Parse file content based on file type"""
        try:
            if file_extension in ['csv', 'txt', 'tsv']:
                return await self._parse_csv_content(file_content, filename, file_extension)
            elif file_extension == 'json':
                return await self._parse_json_content(file_content, filename)
            elif file_extension in ['xlsx', 'xls']:
                return await self._parse_excel_content(file_content, filename)
            else:
                raise ValueError(f"Unsupported file type: {file_extension}")
                
        except Exception as e:
            logger.error(f"Failed to parse {filename}: {e}")
            raise
    
    async def _parse_csv_content(self, 
                                file_content: bytes, 
                                filename: str, 
                                file_extension: str) -> Dict[str, Any]:
        """Parse CSV/TSV content using streaming processor"""
        try:
            # Convert bytes to string
            content_str = file_content.decode('utf-8')
            
            # Use streaming processor
            result = await self.streaming_processor.process_csv_stream(
                content_str, 
                filename
            )
            
            # Add file type information
            result['metadata']['file_type'] = file_extension
            result['metadata']['parsing_method'] = 'streaming_csv'
            
            return result
            
        except UnicodeDecodeError:
            # Try different encodings
            for encoding in ['latin-1', 'cp1252', 'iso-8859-1']:
                try:
                    content_str = file_content.decode(encoding)
                    result = await self.streaming_processor.process_csv_stream(
                        content_str, 
                        filename
                    )
                    result['metadata']['file_type'] = file_extension
                    result['metadata']['encoding'] = encoding
                    result['metadata']['parsing_method'] = 'streaming_csv'
                    return result
                except UnicodeDecodeError:
                    continue
            
            raise ValueError("Unable to decode file with any supported encoding")
    
    async def _parse_json_content(self, 
                                 file_content: bytes, 
                                 filename: str) -> Dict[str, Any]:
        """Parse JSON content using streaming processor"""
        try:
            # Convert bytes to string
            content_str = file_content.decode('utf-8')
            
            # Use streaming processor
            result = await self.streaming_processor.process_json_stream(
                content_str, 
                filename
            )
            
            # Add file type information
            result['metadata']['file_type'] = 'json'
            result['metadata']['parsing_method'] = 'streaming_json'
            
            return result
            
        except Exception as e:
            logger.error(f"JSON parsing failed: {e}")
            raise
    
    async def _parse_excel_content(self, 
                                  file_content: bytes, 
                                  filename: str) -> Dict[str, Any]:
        """Parse Excel content with memory optimization"""
        try:
            # Create BytesIO from content
            excel_buffer = BytesIO(file_content)
            
            # Load workbook with read-only mode for memory efficiency
            workbook = load_workbook(excel_buffer, read_only=True, data_only=True)
            
            # Get first worksheet
            if not workbook.sheetnames:
                raise ValueError("Excel file contains no worksheets")
            
            worksheet = workbook[workbook.sheetnames[0]]
            
            # Read data in chunks to manage memory
            data = []
            headers = None
            row_count = 0
            
            for row in worksheet.iter_rows(values_only=True):
                if headers is None:
                    # First row as headers
                    headers = [str(cell) if cell is not None else f"Column_{i}" 
                              for i, cell in enumerate(row)]
                    continue
                
                # Convert row to dictionary
                row_dict = {}
                for i, cell in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = cell if cell is not None else ""
                
                data.append(row_dict)
                row_count += 1
                
                # Yield control periodically
                if row_count % 1000 == 0:
                    await asyncio.sleep(0)
            
            # Close workbook to free memory
            workbook.close()
            
            # Apply sampling if needed
            sampled_data = self.streaming_processor.data_sampler.smart_sample(data)
            
            # Generate metadata
            metadata = {
                "original_rows": len(data),
                "sampled_rows": len(sampled_data),
                "sampling_ratio": len(sampled_data) / len(data) if data else 0,
                "file_size_bytes": len(file_content),
                "file_size_mb": len(file_content) / (1024 * 1024),
                "filename": filename,
                "file_type": "excel",
                "is_sampled": len(sampled_data) < len(data),
                "parsing_method": "excel_streaming",
                "worksheet_name": workbook.sheetnames[0] if workbook.sheetnames else "unknown"
            }
            
            return {
                "data": sampled_data,
                "metadata": metadata,
                "processing_stats": self.streaming_processor.get_processing_stats()
            }
            
        except Exception as e:
            logger.error(f"Excel parsing failed: {e}")
            raise
    
    def _get_file_extension(self, filename: Optional[str]) -> str:
        """Extract file extension from filename"""
        if not filename:
            return "unknown"
        
        return filename.split('.')[-1].lower() if '.' in filename else "unknown"
    
    def validate_file_size(self, file_size_bytes: int, max_size_mb: int = 100) -> bool:
        """Validate file size against limits"""
        max_size_bytes = max_size_mb * 1024 * 1024
        return file_size_bytes <= max_size_bytes
    
    def estimate_processing_time(self, file_size_mb: float) -> int:
        """Estimate processing time in seconds based on file size"""
        # Rough estimate: 1MB per second for CSV, 2MB per second for JSON/Excel
        base_time = max(5, int(file_size_mb))  # Minimum 5 seconds
        return min(base_time, 300)  # Maximum 5 minutes
    
    async def get_file_preview(self, 
                              file_content: bytes, 
                              filename: str, 
                              preview_rows: int = 5) -> Dict[str, Any]:
        """Get a quick preview of file content without full processing"""
        try:
            file_extension = self._get_file_extension(filename)
            
            if file_extension in ['csv', 'txt', 'tsv']:
                # Quick CSV preview
                content_str = file_content.decode('utf-8')
                lines = content_str.split('\n')[:preview_rows + 1]  # +1 for header
                
                if len(lines) < 2:
                    return {"error": "File has insufficient data for preview"}
                
                # Parse just the preview lines
                import csv
                from io import StringIO
                
                preview_content = '\n'.join(lines)
                csv_reader = csv.DictReader(StringIO(preview_content))
                preview_data = list(csv_reader)
                
                return {
                    "preview_data": preview_data,
                    "total_lines_estimate": len(content_str.split('\n')),
                    "columns": list(preview_data[0].keys()) if preview_data else [],
                    "file_type": file_extension
                }
            
            elif file_extension == 'json':
                # Quick JSON preview
                content_str = file_content.decode('utf-8')
                json_data = json.loads(content_str)
                
                if isinstance(json_data, list):
                    preview_data = json_data[:preview_rows]
                    total_count = len(json_data)
                elif isinstance(json_data, dict) and 'data' in json_data:
                    preview_data = json_data['data'][:preview_rows]
                    total_count = len(json_data['data'])
                else:
                    preview_data = [json_data]
                    total_count = 1
                
                return {
                    "preview_data": preview_data,
                    "total_rows_estimate": total_count,
                    "columns": list(preview_data[0].keys()) if preview_data else [],
                    "file_type": file_extension
                }
            
            else:
                return {"error": f"Preview not supported for {file_extension} files"}
                
        except Exception as e:
            logger.error(f"Failed to generate preview for {filename}: {e}")
            return {"error": f"Preview generation failed: {str(e)}"}