"""Tests for app/utils/enhanced_file_parser.py"""

import asyncio
import json
import pytest
from io import BytesIO
from unittest.mock import AsyncMock, MagicMock, patch

from app.utils.enhanced_file_parser import EnhancedFileParser
from app.utils.memory_manager import RequestPriority


@pytest.fixture
def parser():
    return EnhancedFileParser(chunk_size=100, memory_limit_mb=256, sample_size=50)


def make_upload_file(content: bytes, filename: str) -> MagicMock:
    """Create a mock UploadFile."""
    f = MagicMock()
    f.filename = filename
    f.read = AsyncMock(return_value=content)
    return f


# ── _get_file_extension ───────────────────────────────────────────────────────


def test_get_file_extension_csv(parser):
    assert parser._get_file_extension("data.csv") == "csv"


def test_get_file_extension_xlsx(parser):
    assert parser._get_file_extension("report.XLSX") == "xlsx"


def test_get_file_extension_no_dot(parser):
    assert parser._get_file_extension("nodotfile") == "unknown"


def test_get_file_extension_none(parser):
    assert parser._get_file_extension(None) == "unknown"


def test_get_file_extension_empty(parser):
    assert parser._get_file_extension("") == "unknown"


# ── validate_file_size ────────────────────────────────────────────────────────


def test_validate_file_size_ok(parser):
    assert parser.validate_file_size(1024 * 1024) is True  # 1 MB < 100 MB


def test_validate_file_size_too_big(parser):
    assert parser.validate_file_size(200 * 1024 * 1024) is False  # 200 MB > 100 MB


def test_validate_file_size_exact_limit(parser):
    assert parser.validate_file_size(100 * 1024 * 1024) is True


def test_validate_file_size_custom_limit(parser):
    assert parser.validate_file_size(5 * 1024 * 1024, max_size_mb=4) is False


# ── estimate_processing_time ──────────────────────────────────────────────────


def test_estimate_processing_time_small(parser):
    t = parser.estimate_processing_time(0.5)
    assert t == 5  # minimum


def test_estimate_processing_time_normal(parser):
    t = parser.estimate_processing_time(30.0)
    assert t == 30


def test_estimate_processing_time_cap(parser):
    t = parser.estimate_processing_time(500.0)
    assert t == 300  # max 5 minutes


# ── _parse_file_content ───────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_parse_csv_content(parser):
    csv_bytes = b"a,b,c\n1,2,3\n4,5,6\n"
    result = await parser._parse_file_content(csv_bytes, "test.csv", "csv")
    assert "data" in result
    assert result["metadata"]["file_type"] == "csv"
    assert result["metadata"]["parsing_method"] == "streaming_csv"


@pytest.mark.asyncio
async def test_parse_tsv_content(parser):
    tsv_bytes = b"a\tb\n1\t2\n3\t4\n"
    result = await parser._parse_file_content(tsv_bytes, "test.tsv", "tsv")
    assert result["metadata"]["file_type"] == "tsv"


@pytest.mark.asyncio
async def test_parse_json_content(parser):
    data = [{"x": 1}, {"x": 2}]
    json_bytes = json.dumps(data).encode()
    result = await parser._parse_file_content(json_bytes, "test.json", "json")
    assert result["metadata"]["file_type"] == "json"
    assert result["metadata"]["parsing_method"] == "streaming_json"


@pytest.mark.asyncio
async def test_parse_json_content_exception_reraises(parser):
    """Exercise lines 178-180 — exception in _parse_json_content."""
    with patch.object(parser.streaming_processor, "process_json_stream", side_effect=RuntimeError("fail")):
        with pytest.raises(RuntimeError):
            await parser._parse_json_content(b'[{"x":1}]', "test.json")


@pytest.mark.asyncio
async def test_parse_xlsx_via_parse_file_content(parser):
    """Exercise line 111 — _parse_file_content dispatching to _parse_excel_content."""
    import openpyxl
    from io import BytesIO
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["col"])
    ws.append(["val"])
    buf = BytesIO()
    wb.save(buf)
    result = await parser._parse_file_content(buf.getvalue(), "test.xlsx", "xlsx")
    assert result["metadata"]["file_type"] == "excel"


@pytest.mark.asyncio
async def test_parse_xls_via_parse_file_content(parser):
    """Exercise line 111 for 'xls' extension path."""
    import openpyxl
    from io import BytesIO
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["col"])
    buf = BytesIO()
    wb.save(buf)
    result = await parser._parse_file_content(buf.getvalue(), "test.xls", "xls")
    assert "data" in result


@pytest.mark.asyncio
async def test_parse_unsupported_extension(parser):
    with pytest.raises(Exception, match="Unsupported"):
        await parser._parse_file_content(b"data", "test.parquet", "parquet")


# ── _parse_csv_content – encoding fallback ───────────────────────────────────


@pytest.mark.asyncio
async def test_parse_csv_latin1_encoding(parser):
    # latin-1 encoded content
    csv_bytes = "name,city\nJosé,México\n".encode("latin-1")
    result = await parser._parse_csv_content(csv_bytes, "test.csv", "csv")
    assert result["metadata"]["encoding"] == "latin-1"


@pytest.mark.asyncio
async def test_parse_csv_undecodable_raises(parser):
    bad_bytes = b"\x80\x81\x82"  # Invalid in utf-8 and all fallback encodings mock
    with patch.object(parser.streaming_processor, "process_csv_stream", side_effect=UnicodeDecodeError("utf-8", b"", 0, 1, "err")):
        with pytest.raises(Exception):
            await parser._parse_csv_content(bad_bytes, "test.csv", "csv")


# ── _parse_excel_content ──────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_parse_excel_content(parser):
    """Create a minimal real xlsx in-memory."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["name", "score"])
    ws.append(["Alice", 95])
    ws.append(["Bob", 88])
    buf = BytesIO()
    wb.save(buf)
    excel_bytes = buf.getvalue()

    result = await parser._parse_excel_content(excel_bytes, "test.xlsx")
    assert "data" in result
    assert len(result["data"]) == 2
    assert result["metadata"]["file_type"] == "excel"


@pytest.mark.asyncio
async def test_parse_excel_content_over_1000_rows(parser):
    """Exercise asyncio.sleep(0) every 1000 rows (line 222)."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["val"])
    for i in range(1001):
        ws.append([i])
    buf = BytesIO()
    wb.save(buf)
    result = await parser._parse_excel_content(buf.getvalue(), "big.xlsx")
    assert result["metadata"]["original_rows"] == 1001


@pytest.mark.asyncio
async def test_parse_excel_no_worksheets(parser):
    from unittest.mock import patch as mpatch
    mock_wb = MagicMock()
    mock_wb.sheetnames = []
    with mpatch("app.utils.enhanced_file_parser.load_workbook", return_value=mock_wb):
        with pytest.raises(ValueError, match="no worksheets"):
            await parser._parse_excel_content(b"fake", "test.xlsx")


# ── get_file_preview ──────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_get_file_preview_csv(parser):
    csv_bytes = b"name,age\nAlice,30\nBob,25\nCarol,35\n"
    result = await parser.get_file_preview(csv_bytes, "test.csv", preview_rows=2)
    assert "preview_data" in result
    assert len(result["preview_data"]) == 2
    assert result["file_type"] == "csv"


@pytest.mark.asyncio
async def test_get_file_preview_csv_too_short(parser):
    # A file with only a header row (no data) and no trailing newline → 1 line → error
    csv_bytes = b"name"
    result = await parser.get_file_preview(csv_bytes, "test.csv")
    assert "error" in result


@pytest.mark.asyncio
async def test_get_file_preview_json_list(parser):
    data = [{"x": i} for i in range(10)]
    result = await parser.get_file_preview(json.dumps(data).encode(), "test.json", preview_rows=3)
    assert "preview_data" in result
    assert len(result["preview_data"]) == 3


@pytest.mark.asyncio
async def test_get_file_preview_json_dict_with_data(parser):
    payload = {"data": [{"v": i} for i in range(8)]}
    result = await parser.get_file_preview(json.dumps(payload).encode(), "test.json", preview_rows=4)
    assert len(result["preview_data"]) == 4


@pytest.mark.asyncio
async def test_get_file_preview_json_plain_dict(parser):
    payload = {"key": "value"}
    result = await parser.get_file_preview(json.dumps(payload).encode(), "test.json")
    assert result["total_rows_estimate"] == 1


@pytest.mark.asyncio
async def test_get_file_preview_unsupported(parser):
    result = await parser.get_file_preview(b"data", "test.parquet")
    assert "error" in result


@pytest.mark.asyncio
async def test_get_file_preview_exception_returns_error(parser):
    result = await parser.get_file_preview(b"\xff\xfe", "bad.csv")
    assert "error" in result


# ── parse_file (integration through memory manager mock) ─────────────────────


@pytest.mark.asyncio
async def test_parse_file_csv_immediate_execution(parser):
    """queue_request executes callback immediately (memory available path)."""
    csv_bytes = b"col1,col2\n1,2\n3,4\n"
    upload = make_upload_file(csv_bytes, "data.csv")

    async def execute_immediately(request_id, callback, **kwargs):
        await callback()
        return True

    with patch.object(parser.memory_manager, "queue_request", side_effect=execute_immediately):
        result = await parser.parse_file(upload, "req-001")
    assert "data" in result


@pytest.mark.asyncio
async def test_parse_file_csv_queued_execution(parser):
    """queue_request queues without executing; polling loop waits for background execution."""
    csv_bytes = b"col1,col2\n1,2\n3,4\n"
    upload = make_upload_file(csv_bytes, "data.csv")
    captured_callback = {}

    async def capture_callback(request_id, callback, **kwargs):
        captured_callback["fn"] = callback
        return True  # queued but not executed yet

    async def run_queued():
        await asyncio.sleep(0.05)
        await captured_callback["fn"]()

    with patch.object(parser.memory_manager, "queue_request", side_effect=capture_callback):
        bg = asyncio.create_task(run_queued())
        result = await parser.parse_file(upload, "req-001b")
        await bg
    assert "data" in result


@pytest.mark.asyncio
async def test_parse_file_queued_callback_error_propagates(parser):
    """Error from callback in queued path is raised immediately, not after 300s timeout."""
    upload = make_upload_file(b"a,b\n1,2\n", "data.csv")
    captured_callback = {}

    async def capture_callback(request_id, callback, **kwargs):
        captured_callback["fn"] = callback
        return True

    async def run_failing_callback():
        await asyncio.sleep(0.05)
        try:
            await captured_callback["fn"]()
        except Exception:
            pass  # _process_queue swallows errors; result_holder['error'] is set

    with patch.object(parser.memory_manager, "queue_request", side_effect=capture_callback):
        with patch.object(parser, "_parse_file_content", side_effect=RuntimeError("parse failed")):
            bg = asyncio.create_task(run_failing_callback())
            with pytest.raises(RuntimeError, match="parse failed"):
                await parser.parse_file(upload, "req-err")
            await bg


@pytest.mark.asyncio
async def test_parse_file_immediate_execution_error_not_masked_as_overload(parser):
    """Parsing error during immediate execution raises the real error, not 'overloaded'."""
    upload = make_upload_file(b"a,b\n1,2\n", "data.csv")

    async def execute_with_error(request_id, callback, **kwargs):
        try:
            await callback()
        except Exception:
            pass
        return False  # queue_request returns False on execution failure

    with patch.object(parser.memory_manager, "queue_request", side_effect=execute_with_error):
        with patch.object(parser, "_parse_file_content", side_effect=RuntimeError("parse failed")):
            with pytest.raises(RuntimeError, match="parse failed"):
                await parser.parse_file(upload, "req-masked")


@pytest.mark.asyncio
async def test_parse_file_cancellation_cleans_up(parser):
    """CancelledError during polling cancels the queued request."""
    upload = make_upload_file(b"a,b\n1,2\n", "data.csv")

    async def never_execute(request_id, callback, **kwargs):
        return True  # queued, never run

    with patch.object(parser.memory_manager, "queue_request", side_effect=never_execute):
        task = asyncio.create_task(parser.parse_file(upload, "cancel-poll"))
        await asyncio.sleep(0.05)
        task.cancel()
        with pytest.raises(asyncio.CancelledError):
            await task

    assert parser.memory_manager.cancel_request("cancel-poll") is False  # already cancelled


@pytest.mark.asyncio
async def test_parse_file_queue_rejected(parser):
    csv_bytes = b"a,b\n1,2\n"
    upload = make_upload_file(csv_bytes, "data.csv")
    with patch.object(parser.memory_manager, "queue_request", new_callable=AsyncMock, return_value=False):
        with pytest.raises(Exception, match="overloaded"):
            await parser.parse_file(upload, "req-002")
